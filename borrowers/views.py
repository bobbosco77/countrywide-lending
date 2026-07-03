from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.db.models import Q, Sum
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required

from reportlab.pdfgen import canvas
from datetime import date, timedelta
from decimal import Decimal
from functools import wraps
import json
import openpyxl
from .pdf_reports import (
    loan_register_pdf,
    payment_register_pdf,
    loan_statement_pdf,
    borrowers_pdf,
    defaulters_pdf,
    weekly_collection_pdf,
    monthly_financial_report_pdf,
)

from .excel_reports import (
    loan_register_excel,
    payment_register_excel,
)

from .forms import BorrowerForm
from .models import (
    Borrower,
    Loan,
    RepaymentSchedule,
    Payment,
    CompanySettings,
)
from .utils import log_action
from .permissions import (
    is_cashier,
    is_loan_officer,
    is_manager,
    is_auditor,
)


# ==================================================
# ROLE DECORATOR
# ==================================================

def role_required(check):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')

            if not check(request.user):
                return HttpResponse("Access Denied", status=403)

            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator


# ==================================================
# ROLE HELPERS
# ==================================================

def is_manager_or_superuser(user):
    return user.is_superuser or is_manager(user)


# ==================================================
# AUTH
# ==================================================

def user_login(request):

    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user:

            login(request, user)

            messages.success(
                request,
                f"Welcome back, {user.get_full_name() or user.username}!"
            )

            next_url = request.GET.get("next")
            if next_url:
                return redirect(next_url)

            if user.is_superuser or is_manager(user):
                return redirect("dashboard")

            elif is_cashier(user):
                return redirect("borrower_list")

            elif is_auditor(user):
                return redirect("defaulters_report")

            elif is_loan_officer(user):
                return redirect("register")

            return redirect("dashboard")

        messages.error(request, "Invalid username or password.")

    return render(request, "login.html")

# ==================================================
# BORROWER
# ==================================================

@login_required
@role_required(
    lambda u:
    is_manager_or_superuser(u)
    or is_cashier(u)
    or is_loan_officer(u)
)
def register_borrower(request):

    form = BorrowerForm(
        request.POST or None,
        request.FILES or None
    )

    if request.method == 'POST' and form.is_valid():

        borrower = form.save()

        log_action(
            request.user,
            'CREATE',
            borrower,
            f"Borrower {borrower} created"
        )

        # Loan Officer returns to registration page
        if is_loan_officer(request.user):
            messages.success(
                request,
                "Borrower registered successfully."
            )
            return redirect('register')

        # Cashier and Manager go to borrower list
        return redirect('borrower_list')

    return render(
        request,
        'register.html',
        {'form': form}
    )

@login_required
@role_required(
    lambda u:
    is_manager_or_superuser(u)
    or is_cashier(u)
    or is_auditor(u)
)
def borrower_list(request):

    borrowers = Borrower.objects.all()
    query = request.GET.get('q')

    if query:
        borrowers = borrowers.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(phone__icontains=query) |
            Q(nin__icontains=query) |
            Q(bvn__icontains=query)
        )

    return render(
        request,
        'borrowers/borrower_list.html',
        {
            'borrowers': borrowers,
            'query': query
        }
    )


@login_required
def borrower_profile(request, borrower_id):
    borrower = get_object_or_404(Borrower, id=borrower_id)

    loans = (
        Loan.objects
        .filter(borrower=borrower)
        .order_by("-start_date")
    )

    payments = (
        Payment.objects
        .filter(loan__borrower=borrower)
        .select_related("loan")
        .order_by("-payment_date")
    )

    schedules = (
        RepaymentSchedule.objects
        .filter(loan__borrower=borrower)
    )

    total_borrowed = sum(
        loan.loan_amount
        for loan in loans
    )

    total_balance = sum(
        loan.balance
        for loan in loans
    )

    return render(
        request,
        "borrowers/profile.html",
        {
            "borrower": borrower,
            "loans": loans,
            "payments": payments,
            "schedules": schedules,
            "total_borrowed": total_borrowed,
            "loan_balance": total_balance,
        }
    )


# ==================================================
# LOANS
# ==================================================

@login_required
@role_required(lambda u: is_manager_or_superuser(u) or is_loan_officer(u))
def create_loan(request):

    borrowers = Borrower.objects.order_by("first_name", "last_name")

    if request.method == "POST":

        borrower = get_object_or_404(
            Borrower,
            id=request.POST.get("borrower_id")
        )

        amount = Decimal(request.POST.get("loan_amount"))
        rate = Decimal(request.POST.get("interest_rate"))
        weeks = int(request.POST['duration_weeks'])

        loan = Loan.objects.create(
            borrower=borrower,
            loan_amount=amount,
            interest_rate=rate,
            duration_weeks=weeks,
            status="active"
        )

        # Generate repayment schedule
        loan.generate_repayment_schedule()

        log_action(
            request.user,
            "LOAN",
            loan,
            f"Loan #{loan.id} created for {borrower.first_name} {borrower.last_name}"
        )

        messages.success(
            request,
            "Loan created successfully."
        )

        return redirect("loan_form")

    return render(
        request,
        "loan_form.html",
        {
            "borrowers": borrowers,
        }
    )

# ==================================================
# LOAN REGISTER
# ==================================================

@login_required
@role_required(
    lambda u:
    is_manager_or_superuser(u)
    or is_cashier(u)
    or is_loan_officer(u)
    or is_auditor(u)
)
def loan_register(request):

    loans = (
        Loan.objects
        .select_related("borrower")
        .order_by("-start_date")
    )

    # -----------------------------
    # Search
    # -----------------------------
    search = request.GET.get("search")

    if search:
        loans = loans.filter(
            Q(borrower__first_name__icontains=search) |
            Q(borrower__last_name__icontains=search) |
            Q(borrower__phone__icontains=search) |
            Q(borrower__nin__icontains=search) |
            Q(borrower__bvn__icontains=search)
        )

    # -----------------------------
    # Status Filter
    # -----------------------------
    status = request.GET.get("status")

    if status:
        loans = loans.filter(status=status)

    # -----------------------------
    # Date Filter
    # -----------------------------
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    if from_date:
        loans = loans.filter(start_date__gte=from_date)

    if to_date:
        loans = loans.filter(start_date__lte=to_date)

    # -----------------------------
    # Dashboard Statistics
    # -----------------------------
    total_loans = loans.count()

    active_loans = loans.filter(status="active").count()

    closed_loans = loans.filter(status="closed").count()

    total_portfolio = (
        loans.aggregate(
            total=Sum("loan_amount")
        )["total"] or Decimal("0.00")
    )

    total_paid = sum(
        loan.total_paid
        for loan in loans
    )

    outstanding = sum(
        loan.balance
        for loan in loans
    )

    context = {
        "loans": loans,
        "active_loans": loans.filter(status="active").count(),
        "total_loans": total_loans,
        "active_loans": active_loans,
        "closed_loans": closed_loans,
        "total_portfolio": total_portfolio,
        "total_paid": total_paid,
        "outstanding": outstanding,
        "search": search,
        "status": status,
        "from_date": from_date,
        "to_date": to_date,
    }

    return render(
        request,
        "borrowers/loan_register.html",
        context,
    )


# ==================================================
# PAYMENT
# ==================================================

@login_required
@role_required(lambda u: is_manager_or_superuser(u) or is_cashier(u))
def make_payment(request):
    loans = Loan.objects.filter(status__in=['active', 'approved'])

    if request.method == 'POST':
        loan = get_object_or_404(Loan, id=request.POST['loan_id'])
        amount = Decimal(request.POST['amount_paid'])

        if amount <= 0:
            messages.error(request, "Payment must be greater than zero")
            return redirect('make_payment')

        payment = Payment.objects.create(
            loan=loan,
            amount_paid=amount,
            method=request.POST.get('method', 'cash')
        )


        return redirect('payment_receipt', payment.id)
    return render(request, 'borrowers/make_payment.html', {'loans': loans})

@login_required
@role_required(lambda u: is_manager_or_superuser(u) or is_cashier(u))
def payment_receipt(request, payment_id):

    payment = get_object_or_404(
        Payment.objects.select_related(
            "loan",
            "loan__borrower"
        ),
        id=payment_id
    )

    loan = payment.loan

    context = {
        "payment": payment,
        "loan": loan,
        "borrower": loan.borrower,
        "total_paid": loan.total_paid,
        "balance": loan.balance,
    }

    return render(
        request,
        "borrowers/receipt.html",
        context
    )



# ==================================================
# DASHBOARD
# ==================================================

@login_required
@role_required(is_manager_or_superuser)
def dashboard(request):

    recent_loans = (
        Loan.objects
        .select_related("borrower")
        .order_by("-id")[:5]
    )

    recent_payments = (
        Payment.objects
        .select_related("loan", "loan__borrower")
        .order_by("-id")[:5]
    )

    total_portfolio = Loan.objects.aggregate(
        total=Sum("loan_amount")
    )["total"] or 0

    total_collections = Payment.objects.aggregate(
        total=Sum("amount_paid")
    )["total"] or 0

    outstanding_balance = sum(
        loan.balance for loan in Loan.objects.all()
    )

    today_collections = Payment.objects.filter(
        payment_date=date.today()
    ).aggregate(
        total=Sum("amount_paid")
    )["total"] or 0

    context = {

        "total_borrowers": Borrower.objects.count(),

        "active_loans": Loan.objects.filter(
            status="active"
        ).count(),

        "closed_loans": Loan.objects.filter(
            status="closed"
        ).count(),

        "pending_loans": Loan.objects.filter(
            status="pending"
        ).count(),

        "total_portfolio": total_portfolio,

        "total_repayments": total_collections,

        "loan_balance": outstanding_balance,

        "today_collections": today_collections,

        "recent_loans": recent_loans,

        "recent_payments": recent_payments,

    }

    return render(
        request,
        "borrowers/dashboard.html",
        context,
    )
# ==================================================
# LOGIN / AUDIT / REPORTS
# ==================================================

@login_required
def loan_detail(request, loan_id):

    loan = get_object_or_404(
        Loan.objects.select_related("borrower"),
        id=loan_id
    )

    schedules = (
        RepaymentSchedule.objects
        .filter(loan=loan)
        .order_by("installment_number")
    )

    payments = (
        Payment.objects
        .filter(loan=loan)
        .order_by("-payment_date")
    )

    total_paid = loan.total_paid
    total_repayment = loan.total_repayment()
    balance = loan.balance

    total_installments = schedules.count()
    paid_installments = schedules.filter(status="paid").count()

    # Progress based on amount paid
    progress = 0

    if total_repayment > 0:
        progress = round(
            float((total_paid / total_repayment) * 100),
            1
        )

    # Never exceed 100%
    progress = min(progress, 100)

    context = {
        "loan": loan,
        "schedules": schedules,
        "payments": payments,
        "total_paid": total_paid,
        "total_repayment": total_repayment,
        "balance": balance,
        "total_installments": total_installments,
        "paid_installments": paid_installments,
        "progress": progress,
    }

    return render(
        request,
        "borrowers/loan_detail.html",
        context,
    )


@login_required
def user_logout(request):
    logout(request)
    return redirect('login')

@login_required
@role_required(lambda u: is_manager_or_superuser(u) or is_auditor(u))
def defaulters_report(request):

    overdue = RepaymentSchedule.objects.select_related(
        'loan',
        'loan__borrower'
    ).filter(
        status='pending',
        due_date__lt=date.today()
    )

    search = request.GET.get('search')

    if search:
        overdue = overdue.filter(
            Q(loan__borrower__first_name__icontains=search) |
            Q(loan__borrower__last_name__icontains=search) |
            Q(loan__borrower__phone__icontains=search)
        )

    for item in overdue:
        item.days_overdue = (date.today() - item.due_date).days

    chart_labels = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]
    chart_data = [0,0,0,0,0,0,0]

    context = {
        "defaulters": overdue,
        "chart_labels": json.dumps(chart_labels),
        "chart_data": json.dumps(chart_data),
        "officers": [],   # keeps your template working
    }

    return render(
        request,
        "borrowers/defaulters.html",
        context
    )

@login_required
@role_required(lambda u: is_manager_or_superuser(u))
def payment_register_pdf_view(request):
    return payment_register_pdf(request)

@login_required
@role_required(lambda u: is_manager_or_superuser(u))
def loan_register_pdf_view(request):
    return loan_register_pdf(request)

@login_required
@role_required(lambda u: is_manager_or_superuser(u))
def loan_statement_pdf_view(request):
    return loan_statement_pdf(request)


@login_required
@role_required(lambda u: is_manager_or_superuser(u))
def borrowers_pdf_view(request):
    return borrowers_pdf(request)

@login_required
@role_required(lambda u: is_manager_or_superuser(u))
def defaulters_pdf_view(request):
    return defaulters_pdf(request)


@login_required
@role_required(lambda u: is_manager_or_superuser(u))
def weekly_collection_pdf_view(request):
    return weekly_collection_pdf(request)


@login_required
@role_required(lambda u: is_manager_or_superuser(u))
def monthly_financial_report_pdf_view(request):
    return monthly_financial_report_pdf(request)


@login_required
@role_required(lambda u: is_manager_or_superuser(u))
def loan_register_excel_view(request):
    return loan_register_excel(request)


@login_required
@role_required(lambda u: is_manager_or_superuser(u))
def payment_register_excel_view(request):
    return payment_register_excel(request)


@login_required
@role_required(lambda u: is_manager_or_superuser(u) or is_auditor(u))
def export_defaulters_excel(request):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defaulters"

    ws.append(["Borrower", "Phone", "Amount Due", "Due Date"])

    defaulters = RepaymentSchedule.objects.filter(
        due_date__lt=date.today()
    ).exclude(status='paid')

    for item in defaulters:
        ws.append([
            f"{item.loan.borrower.first_name} {item.loan.borrower.last_name}",
            item.loan.borrower.phone,
            item.amount_due,
            item.due_date,
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=defaulters.xlsx'

    wb.save(response)
    return response

@login_required
@role_required(lambda u: is_manager_or_superuser(u) or is_auditor(u))
def export_defaulters_pdf(request):

    overdue = RepaymentSchedule.objects.filter(
        status='pending',
        due_date__lt=date.today()
    )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="defaulters.pdf"'

    p = canvas.Canvas(response)

    p.drawString(200, 800, "DEFAULTERS REPORT")

    y = 760
    for item in overdue:
        p.drawString(
            50, y,
            f"{item.loan.borrower.first_name} {item.loan.borrower.last_name} - ₦{item.amount_due}"
        )
        y -= 20

        if y < 50:
            p.showPage()
            y = 800

    p.save()
    return response

@login_required
@role_required(is_manager_or_superuser)
def export_loans_pdf(request):

    loans = Loan.objects.select_related("borrower").order_by("-start_date")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        'attachment; filename="Loan_Register.pdf"'
    )

    p = canvas.Canvas(response)

    width, height = 595, 842

    # -----------------------------
    # Header
    # -----------------------------
    p.setFont("Helvetica-Bold", 16)
    company = CompanySettings.objects.first()
    p.drawString(40, height - 40, company.company_name)

    p.setFont("Helvetica", 11)
    p.drawString(40, height - 60, "Loan Register Report")

    p.drawRightString(
        width - 40,
        height - 60,
        date.today().strftime("%d %B %Y")
    )

    y = height - 95

    # -----------------------------
    # Table Header
    # -----------------------------
    p.setFont("Helvetica-Bold", 9)

    p.drawString(40, y, "Borrower")
    p.drawString(170, y, "Loan")
    p.drawString(250, y, "Interest")
    p.drawString(310, y, "Weeks")
    p.drawString(360, y, "Balance")
    p.drawString(450, y, "Status")

    y -= 18

    p.line(40, y + 8, width - 40, y + 8)

    total_portfolio = Decimal("0.00")
    total_balance = Decimal("0.00")

    p.setFont("Helvetica", 8)

    # -----------------------------
    # Loan Records
    # -----------------------------
    for loan in loans:

        if y < 70:

            p.showPage()

            y = height - 50

            p.setFont("Helvetica-Bold", 9)

            p.drawString(40, y, "Borrower")
            p.drawString(170, y, "Loan")
            p.drawString(250, y, "Interest")
            p.drawString(310, y, "Weeks")
            p.drawString(360, y, "Balance")
            p.drawString(450, y, "Status")

            y -= 18

            p.line(40, y + 8, width - 40, y + 8)

            p.setFont("Helvetica", 8)

        borrower = (
            f"{loan.borrower.first_name} "
            f"{loan.borrower.last_name}"
        )

        balance = loan.balance

        total_portfolio += loan.loan_amount
        total_balance += balance

        p.drawString(40, y, borrower[:24])

        p.drawRightString(
            240,
            y,
            f"₦{loan.loan_amount:,.2f}"
        )

        p.drawString(
            255,
            y,
            f"{loan.interest_rate}%"
        )

        p.drawString(
            320,
            y,
            str(loan.duration_weeks)
        )

        p.drawRightString(
            435,
            y,
            f"₦{balance:,.2f}"
        )

        p.drawString(
            455,
            y,
            loan.status.title()
        )

        y -= 18

    # -----------------------------
    # Totals
    # -----------------------------
    y -= 15

    p.line(40, y + 8, width - 40, y + 8)

    p.setFont("Helvetica-Bold", 10)

    p.drawString(40, y - 10, "Total Portfolio")

    p.drawRightString(
        250,
        y - 10,
        f"₦{total_portfolio:,.2f}"
    )

    p.drawString(300, y - 10, "Outstanding")

    p.drawRightString(
        width - 40,
        y - 10,
        f"₦{total_balance:,.2f}"
    )

    # -----------------------------
    # Footer
    # -----------------------------
    p.setFont("Helvetica", 8)

    p.drawCentredString(
        width / 2,
        20,
        "Generated by CountryWide Lending Management System"
    )

    p.save()

    return response

@login_required
@role_required(is_manager_or_superuser)
def export_payments_pdf(request):

    payments = (
        Payment.objects
        .select_related("loan", "loan__borrower")
        .order_by("-payment_date", "-id")
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        'attachment; filename="Payment_Register.pdf"'
    )

    p = canvas.Canvas(response)

    width, height = 595, 842

   # ---------------------------------------
    # COMPANY HEADER
    # ---------------------------------------

    company = CompanySettings.objects.first()

    company_name = (
        company.company_name
        if company
        else "COUNTRYWIDE LENDING & SERVICES"
    )

    company_phone = company.phone if company else ""
    company_email = company.email if company else ""
    company_address = company.address if company else ""

    p.setFont("Helvetica-Bold", 16)
    p.drawString(
        40,
        height - 40,
        company_name
    )

    p.setFont("Helvetica", 9)

    if company_address:
        p.drawString(
            40,
            height - 58,
            company_address
        )

    if company_phone:
        p.drawString(
            40,
            height - 72,
            f"Phone: {company_phone}"
        )

    if company_email:
        p.drawString(
            40,
            height - 86,
            f"Email: {company_email}"
        )

    p.setFont("Helvetica-Bold", 12)
    p.drawCentredString(
        width / 2,
        height - 105,
        "PAYMENT REGISTER"
    )

    p.setFont("Helvetica", 9)
    p.drawRightString(
        width - 40,
        height - 105,
        date.today().strftime("%d %B %Y")
    )

    y = height - 130

    # ---------------------------------------
    # TABLE HEADER
    # ---------------------------------------

    p.setFont("Helvetica-Bold", 8)

    p.drawString(40, y, "Receipt")
    p.drawString(115, y, "Date")
    p.drawString(185, y, "Borrower")
    p.drawString(335, y, "Method")
    p.drawString(415, y, "Loan")
    p.drawString(485, y, "Amount")

    y -= 18

    p.line(40, y + 8, width - 40, y + 8)

    p.setFont("Helvetica", 8)

    total_payments = Decimal("0.00")
    # ---------------------------------------
    # PAYMENT RECORDS
    # ---------------------------------------

    for payment in payments:

        if y < 70:

            p.showPage()

            y = height - 50

            p.setFont("Helvetica-Bold", 8)

            p.drawString(40, y, "Receipt")
            p.drawString(95, y, "Date")
            p.drawString(170, y, "Borrower")
            p.drawString(315, y, "Method")
            p.drawString(390, y, "Loan ID")
            p.drawString(470, y, "Amount")

            y -= 18

            p.line(40, y + 8, width - 40, y + 8)

            p.setFont("Helvetica", 8)

        borrower = (
            f"{payment.loan.borrower.first_name} "
            f"{payment.loan.borrower.last_name}"
        )

        receipt = f"RCP{payment.id:06d}"

        total_payments += payment.amount_paid

        p.drawString(40, y, receipt)

        p.drawString(
            95,
            y,
            payment.payment_date.strftime("%d/%m/%Y")
        )

        p.drawString(
            170,
            y,
            borrower[:24]
        )

        p.drawString(
            315,
            y,
            payment.get_method_display()
        )

        p.drawString(
            400,
            y,
            str(payment.loan.id)
        )

        p.drawRightString(
            width - 40,
            y,
            f"₦{payment.amount_paid:,.2f}"
        )

        y -= 18

    # ---------------------------------------
    # TOTALS
    # ---------------------------------------

    y -= 15

    p.line(40, y + 8, width - 40, y + 8)

    p.setFont("Helvetica-Bold", 10)

    p.drawString(
        40,
        y - 10,
        "Total Payments Received"
    )

    p.drawRightString(
        width - 40,
        y - 10,
        f"₦{total_payments:,.2f}"
    )

    # ---------------------------------------
    # FOOTER
    # ---------------------------------------

    p.setFont("Helvetica", 8)

    p.drawCentredString(
        width / 2,
        20,
        "Generated by CountryWide Lending Management System"
    )

    p.save()

    return response

@login_required
@role_required(is_manager_or_superuser)
def export_loans_excel(request):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan Register"

    ws.append([
        "Borrower",
        "Phone",
        "Loan Amount",
        "Interest %",
        "Duration (Weeks)",
        "Total Repayment",
        "Total Paid",
        "Outstanding Balance",
        "Status",
        "Start Date",
    ])

    loans = Loan.objects.select_related("borrower").order_by("-start_date")

    for loan in loans:

        ws.append([
            f"{loan.borrower.first_name} {loan.borrower.last_name}",
            loan.borrower.phone,
            float(loan.loan_amount),
            float(loan.interest_rate),
            loan.duration_weeks,
            float(loan.total_repayment()),
            float(loan.total_paid),
            float(loan.balance),
            loan.status.title(),
            loan.start_date.strftime("%d-%m-%Y"),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Loan_Register.xlsx"'
    )

    wb.save(response)

    return response

@login_required
@role_required(is_manager_or_superuser)
def export_payments_excel(request):

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Payment Register"

    ws.append([
        "Receipt No.",
        "Date",
        "Borrower",
        "Phone",
        "Loan ID",
        "Payment Method",
        "Amount Paid",
    ])

    payments = (
        Payment.objects
        .select_related("loan", "loan__borrower")
        .order_by("-payment_date", "-id")
    )

    for payment in payments:

        ws.append([
            f"RCP{payment.id:06d}",
            payment.payment_date.strftime("%d-%m-%Y"),
            f"{payment.loan.borrower.first_name} {payment.loan.borrower.last_name}",
            payment.loan.borrower.phone,
            payment.loan.id,
            payment.get_method_display(),
            float(payment.amount_paid),
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="Payment_Register.xlsx"'
    )

    wb.save(response)

    return response



# ==================================================
# LOAN REPORTS
# ==================================================

@login_required
@role_required(is_manager_or_superuser)
def all_loans_report(request):

    loans = Loan.objects.select_related("borrower").order_by("-start_date")

    search = request.GET.get("search")
    status = request.GET.get("status")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    if search:
        loans = loans.filter(
            Q(borrower__first_name__icontains=search) |
            Q(borrower__last_name__icontains=search) |
            Q(borrower__phone__icontains=search)
        )

    if status:
        loans = loans.filter(status=status)

    if from_date:
        loans = loans.filter(start_date__gte=from_date)

    if to_date:
        loans = loans.filter(start_date__lte=to_date)

    total_portfolio = loans.aggregate(
        total=Sum("loan_amount")
    )["total"] or Decimal("0.00")

    total_balance = sum(
        loan.balance
        for loan in loans
    )

    context = {
        "loans": loans,
        "total_portfolio": total_portfolio,
        "total_balance": total_balance,
        "statuses": Loan.STATUS_CHOICES,
    }

    return render(
        request,
        "borrowers/reports/all_loans_report.html",
        context
    )


# ==================================================
# PAYMENT REPORTS
# ==================================================

@login_required
@role_required(is_manager_or_superuser)
def all_payments_report(request):

    payments = (
        Payment.objects
        .select_related("loan", "loan__borrower")
        .order_by("-payment_date", "-id")
    )

    search = request.GET.get("search")
    method = request.GET.get("method")
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")

    if search:
        payments = payments.filter(
            Q(loan__borrower__first_name__icontains=search) |
            Q(loan__borrower__last_name__icontains=search) |
            Q(loan__borrower__phone__icontains=search)
        )

    if method:
        payments = payments.filter(method=method)

    if from_date:
        payments = payments.filter(payment_date__gte=from_date)

    if to_date:
        payments = payments.filter(payment_date__lte=to_date)

    total_payments = payments.aggregate(
        total=Sum("amount_paid")
    )["total"] or Decimal("0.00")

    context = {
        "payments": payments,
        "total_payments": total_payments,
        "methods": Payment.PAYMENT_METHOD,
    }

    return render(
        request,
        "borrowers/reports/all_payments_report.html",
        context
    )

@login_required
@role_required(is_manager_or_superuser)
def report_center(request):
    return render(
        request,
        "borrowers/report_center.html"
    )